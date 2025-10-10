"""
DISC Assessment Bank Importer

Handles importing DISC assessment data from Excel/CSV files with:
- 12 sheet structure validation
- Bilingual PT/EN field support
- Automatic translation of missing EN fields
- Version tracking and idempotent upserts
"""
import io
import re
import csv
import logging
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime, date

logger = logging.getLogger(__name__)


class DISCImporter:
    """Import DISC assessment bank data from Excel or CSV files."""

    # Expected sheet structure
    EXPECTED_SHEETS = {
        'Itens_Normativos': {'min_rows': 120, 'table': 'disc_items_norm'},
        'Blocos_Ipsativos': {'min_rows': 24, 'table': 'disc_blocks_ipsa'},
        'Cenarios': {'min_rows': 12, 'table': 'disc_scenarios'},
        'Likert_Mapa': {'min_rows': 5, 'table': 'disc_likert_map'},
        'Regras_Score': {'min_rows': 1, 'table': 'disc_rules'},
        'Normas_Thresholds': {'min_rows': 4, 'table': 'disc_thresholds'},
        'Relatorio_Templates': {'min_rows': 1, 'table': 'disc_report_templates'},
        'Palavras_Descritivas': {'min_rows': 12, 'table': 'disc_words'},
        'Guia_Entrevista': {'min_rows': 1, 'table': 'disc_interview'},
        'Validacao_Qualidade': {'min_rows': 1, 'table': 'disc_quality_checks'},
        'Manifesto_Ético': {'min_rows': 1, 'table': 'disc_ethics'},
        'Versao': {'min_rows': 1, 'table': 'disc_versions'},
    }

    def __init__(self, supabase_client, translator=None):
        """
        Initialize importer.

        Args:
            supabase_client: Supabase client instance
            translator: Translation function(text, target_lang) -> str
        """
        self.supabase = supabase_client
        self.translator = translator
        self.errors = []
        self.warnings = []

    def import_from_file(self, file_path: Optional[str] = None,
                        file_buffer: Optional[bytes] = None,
                        version: Optional[str] = None) -> Dict[str, Any]:
        """
        Import DISC data from Excel or CSV file.

        Args:
            file_path: Path to file (if reading from disk)
            file_buffer: File content as bytes (if reading from upload)
            version: Version string (auto-detected if not provided)

        Returns:
            Dict with import results: {ok: bool, version: str, counts: {}, errors: []}
        """
        self.errors = []
        self.warnings = []

        try:
            # Determine file format and load data
            if file_buffer:
                sheets_data = self._load_from_buffer(file_buffer)
            elif file_path:
                sheets_data = self._load_from_path(file_path)
            else:
                raise ValueError("Either file_path or file_buffer must be provided")

            # Validate sheet structure
            validation_result = self._validate_sheets(sheets_data)
            if not validation_result['valid']:
                return {
                    'ok': False,
                    'errors': validation_result['errors'],
                    'warnings': self.warnings
                }

            # Extract version
            if not version:
                version = self._extract_version(sheets_data.get('Versao', []))

            if not version:
                raise ValueError("Version not provided and could not be auto-detected")

            # Create or get version record
            version_ref = self._upsert_version(version, sheets_data.get('Versao', []))

            # Import all data types
            counts = {}
            counts['norm'] = self._import_normative_items(
                version_ref, sheets_data.get('Itens_Normativos', [])
            )
            counts['ipsa'] = self._import_ipsative_blocks(
                version_ref, sheets_data.get('Blocos_Ipsativos', [])
            )
            counts['scenarios'] = self._import_scenarios(
                version_ref, sheets_data.get('Cenarios', [])
            )
            counts['likert'] = self._import_likert_map(
                version_ref, sheets_data.get('Likert_Mapa', [])
            )
            counts['rules'] = self._import_rules(
                version_ref, sheets_data.get('Regras_Score', [])
            )
            counts['thresholds'] = self._import_thresholds(
                version_ref, sheets_data.get('Normas_Thresholds', [])
            )
            counts['templates'] = self._import_report_templates(
                version_ref, sheets_data.get('Relatorio_Templates', [])
            )
            counts['words'] = self._import_words(
                version_ref, sheets_data.get('Palavras_Descritivas', [])
            )
            counts['interview'] = self._import_interview(
                version_ref, sheets_data.get('Guia_Entrevista', [])
            )
            counts['quality'] = self._import_quality_checks(
                version_ref, sheets_data.get('Validacao_Qualidade', [])
            )
            counts['ethics'] = self._import_ethics(
                version_ref, sheets_data.get('Manifesto_Ético', [])
            )

            return {
                'ok': True,
                'version': version,
                'version_ref': version_ref,
                'counts': counts,
                'warnings': self.warnings,
                'errors': []
            }

        except Exception as e:
            logger.exception("Error importing DISC data")
            return {
                'ok': False,
                'errors': [str(e)] + self.errors,
                'warnings': self.warnings
            }

    def _load_from_buffer(self, buffer: bytes) -> Dict[str, List[Dict]]:
        """Load data from file buffer (XLSX or CSV)."""
        # Try XLSX first
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
            return self._parse_xlsx(workbook)
        except Exception as xlsx_error:
            logger.debug(f"Not XLSX format: {xlsx_error}")

        # Try CSV
        try:
            return self._parse_csv_bundle(buffer)
        except Exception as csv_error:
            logger.debug(f"Not CSV format: {csv_error}")
            raise ValueError("File is neither valid XLSX nor CSV format")

    def _load_from_path(self, path: str) -> Dict[str, List[Dict]]:
        """Load data from file path."""
        with open(path, 'rb') as f:
            return self._load_from_buffer(f.read())

    def _parse_xlsx(self, workbook) -> Dict[str, List[Dict]]:
        """Parse XLSX workbook into dict of sheet data."""
        sheets_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')

            # Parse data rows
            rows = []
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    if header:
                        cell = sheet.cell(row_idx, col_idx)
                        value = cell.value

                        # Normalize value
                        if value is not None:
                            if isinstance(value, str):
                                value = value.strip()
                            row_data[header] = value
                        else:
                            row_data[header] = None

                # Only add non-empty rows
                if any(v is not None and v != '' for v in row_data.values()):
                    rows.append(row_data)

            sheets_data[sheet_name] = rows

        return sheets_data

    def _parse_csv_bundle(self, buffer: bytes) -> Dict[str, List[Dict]]:
        """Parse multiple CSV files from a bundle (not implemented - needs ZIP)."""
        raise NotImplementedError("CSV bundle import not yet implemented")

    def _validate_sheets(self, sheets_data: Dict[str, List[Dict]]) -> Dict[str, Any]:
        """Validate that all required sheets exist with minimum row counts."""
        errors = []

        for sheet_name, config in self.EXPECTED_SHEETS.items():
            if sheet_name not in sheets_data:
                errors.append(f"Missing required sheet: {sheet_name}")
                continue

            row_count = len(sheets_data[sheet_name])
            min_rows = config['min_rows']

            if row_count < min_rows:
                errors.append(
                    f"Sheet '{sheet_name}' has {row_count} rows, expected at least {min_rows}"
                )

        return {
            'valid': len(errors) == 0,
            'errors': errors
        }

    def _extract_version(self, version_rows: List[Dict]) -> Optional[str]:
        """Extract version string from Versao sheet."""
        if not version_rows:
            return None

        first_row = version_rows[0]

        # Try common field names
        for key in ['version', 'versao', 'Version', 'Versao', 'VERSION', 'VERSAO']:
            if key in first_row and first_row[key]:
                return str(first_row[key]).strip()

        return None

    def _upsert_version(self, version: str, version_rows: List[Dict]) -> str:
        """Create or update version record, return version_ref UUID."""
        # Extract metadata
        notes = None
        data_criacao = None

        if version_rows:
            first_row = version_rows[0]
            notes = first_row.get('notes') or first_row.get('notas')

            # Try to extract date
            date_val = first_row.get('data_criacao') or first_row.get('data')
            if date_val:
                if isinstance(date_val, date):
                    data_criacao = date_val.isoformat()
                elif isinstance(date_val, datetime):
                    data_criacao = date_val.date().isoformat()
                elif isinstance(date_val, str):
                    data_criacao = date_val

        # Check if version exists
        result = self.supabase.table('disc_versions').select('id').eq('version', version).execute()

        if result.data:
            # Version exists, return existing ID
            version_ref = result.data[0]['id']
            logger.info(f"Using existing DISC version: {version} (ID: {version_ref})")
        else:
            # Create new version
            insert_data = {
                'version': version,
                'notes': notes,
                'data_criacao': data_criacao
            }
            result = self.supabase.table('disc_versions').insert(insert_data).execute()
            version_ref = result.data[0]['id']
            logger.info(f"Created new DISC version: {version} (ID: {version_ref})")

        return version_ref

    def _translate_if_missing(self, pt_text: Optional[str], en_text: Optional[str]) -> str:
        """Translate PT to EN if EN is missing."""
        if en_text and en_text.strip():
            return en_text.strip()

        if not pt_text or not pt_text.strip():
            return ''

        if self.translator:
            try:
                return self.translator(pt_text.strip(), 'en')
            except Exception as e:
                logger.warning(f"Translation failed: {e}")
                return ''

        return ''

    def _safe_float(self, value: Any, default: float = 0.0) -> float:
        """Safely convert value to float."""
        if value is None or value == '':
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _safe_int(self, value: Any, default: int = 0) -> int:
        """Safely convert value to int."""
        if value is None or value == '':
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default

    def _safe_bool(self, value: Any, default: bool = False) -> bool:
        """Safely convert value to bool."""
        if value is None or value == '':
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            return value.lower() in ('true', 'sim', 'yes', '1', 't', 's', 'y')
        return default

    def _import_normative_items(self, version_ref: str, rows: List[Dict]) -> int:
        """Import normative items (120 expected)."""
        count = 0

        for row in rows:
            item_id = row.get('item_id') or row.get('ItemID') or row.get('ID')
            if not item_id:
                continue

            texto_pt = row.get('texto_pt') or row.get('Texto') or ''
            texto_en = self._translate_if_missing(texto_pt, row.get('texto_en'))

            texto_simples_pt = row.get('texto_simples_pt') or row.get('TextoSimples') or ''
            texto_simples_en = self._translate_if_missing(
                texto_simples_pt, row.get('texto_simples_en')
            )

            data = {
                'version_ref': version_ref,
                'item_id': str(item_id),
                'texto_pt': texto_pt,
                'texto_simples_pt': texto_simples_pt,
                'texto_en': texto_en,
                'texto_simples_en': texto_simples_en,
                'fator_primario': row.get('fator_primario') or row.get('FatorPrimario'),
                'fatores_sec': row.get('fatores_sec') or row.get('FatoresSec'),
                'invertido': self._safe_bool(row.get('invertido')),
                'peso_d': self._safe_float(row.get('peso_d') or row.get('PesoD')),
                'peso_i': self._safe_float(row.get('peso_i') or row.get('PesoI')),
                'peso_s': self._safe_float(row.get('peso_s') or row.get('PesoS')),
                'peso_c': self._safe_float(row.get('peso_c') or row.get('PesoC')),
                'sutil': self._safe_bool(row.get('sutil')),
                'contexto': row.get('contexto'),
                'leitura_nivel': row.get('leitura_nivel') or row.get('LeituraNivel'),
                'versao_str': row.get('versao_str') or row.get('Versao')
            }

            # Upsert
            self.supabase.table('disc_items_norm').upsert(
                data,
                on_conflict='version_ref,item_id'
            ).execute()
            count += 1

        logger.info(f"Imported {count} normative items")
        return count

    def _import_ipsative_blocks(self, version_ref: str, rows: List[Dict]) -> int:
        """Import ipsative blocks (24 expected)."""
        count = 0

        for row in rows:
            bloco_id = row.get('bloco_id') or row.get('BlocoID') or row.get('ID')
            if not bloco_id:
                continue

            data = {
                'version_ref': version_ref,
                'bloco_id': str(bloco_id),
                'frase_a_pt': row.get('frase_a_pt') or row.get('FraseA_PT') or '',
                'frase_a_en': self._translate_if_missing(
                    row.get('frase_a_pt') or row.get('FraseA_PT'),
                    row.get('frase_a_en')
                ),
                'fator_a': row.get('fator_a') or row.get('FatorA'),
                'frase_b_pt': row.get('frase_b_pt') or row.get('FraseB_PT') or '',
                'frase_b_en': self._translate_if_missing(
                    row.get('frase_b_pt') or row.get('FraseB_PT'),
                    row.get('frase_b_en')
                ),
                'fator_b': row.get('fator_b') or row.get('FatorB'),
                'frase_c_pt': row.get('frase_c_pt') or row.get('FraseC_PT') or '',
                'frase_c_en': self._translate_if_missing(
                    row.get('frase_c_pt') or row.get('FraseC_PT'),
                    row.get('frase_c_en')
                ),
                'fator_c': row.get('fator_c') or row.get('FatorC'),
                'frase_d_pt': row.get('frase_d_pt') or row.get('FraseD_PT') or '',
                'frase_d_en': self._translate_if_missing(
                    row.get('frase_d_pt') or row.get('FraseD_PT'),
                    row.get('frase_d_en')
                ),
                'fator_d': row.get('fator_d') or row.get('FatorD'),
                'regra_pt': row.get('regra_pt'),
                'regra_en': self._translate_if_missing(row.get('regra_pt'), row.get('regra_en'))
            }

            self.supabase.table('disc_blocks_ipsa').upsert(
                data,
                on_conflict='version_ref,bloco_id'
            ).execute()
            count += 1

        logger.info(f"Imported {count} ipsative blocks")
        return count

    def _import_scenarios(self, version_ref: str, rows: List[Dict]) -> int:
        """Import scenarios (12 expected)."""
        count = 0

        for row in rows:
            cenario_id = row.get('cenario_id') or row.get('CenarioID') or row.get('ID')
            if not cenario_id:
                continue

            prompt_pt = row.get('prompt_pt') or row.get('Prompt') or ''

            data = {
                'version_ref': version_ref,
                'cenario_id': str(cenario_id),
                'prompt_pt': prompt_pt,
                'prompt_en': self._translate_if_missing(prompt_pt, row.get('prompt_en')),
                'opc_a_pt': row.get('opc_a_pt') or row.get('OpcaoA_PT') or '',
                'opc_a_en': self._translate_if_missing(
                    row.get('opc_a_pt') or row.get('OpcaoA_PT'),
                    row.get('opc_a_en')
                ),
                'fator_a': row.get('fator_a') or row.get('FatorA'),
                'opc_b_pt': row.get('opc_b_pt') or row.get('OpcaoB_PT') or '',
                'opc_b_en': self._translate_if_missing(
                    row.get('opc_b_pt') or row.get('OpcaoB_PT'),
                    row.get('opc_b_en')
                ),
                'fator_b': row.get('fator_b') or row.get('FatorB'),
                'opc_c_pt': row.get('opc_c_pt') or row.get('OpcaoC_PT') or '',
                'opc_c_en': self._translate_if_missing(
                    row.get('opc_c_pt') or row.get('OpcaoC_PT'),
                    row.get('opc_c_en')
                ),
                'fator_c': row.get('fator_c') or row.get('FatorC'),
                'opc_d_pt': row.get('opc_d_pt') or row.get('OpcaoD_PT') or '',
                'opc_d_en': self._translate_if_missing(
                    row.get('opc_d_pt') or row.get('OpcaoD_PT'),
                    row.get('opc_d_en')
                ),
                'fator_d': row.get('fator_d') or row.get('FatorD'),
                'regra_pt': row.get('regra_pt'),
                'regra_en': self._translate_if_missing(row.get('regra_pt'), row.get('regra_en')),
                'contexto': row.get('contexto')
            }

            self.supabase.table('disc_scenarios').upsert(
                data,
                on_conflict='version_ref,cenario_id'
            ).execute()
            count += 1

        logger.info(f"Imported {count} scenarios")
        return count

    def _import_likert_map(self, version_ref: str, rows: List[Dict]) -> int:
        """Import Likert scale mapping (5 expected)."""
        count = 0

        for row in rows:
            resposta = self._safe_int(row.get('resposta') or row.get('Resposta'))
            if resposta < 1 or resposta > 5:
                continue

            ancora_pt = row.get('ancora_pt') or row.get('Ancora') or ''

            data = {
                'version_ref': version_ref,
                'resposta': resposta,
                'ancora_pt': ancora_pt,
                'ancora_en': self._translate_if_missing(ancora_pt, row.get('ancora_en')),
                'escore_d': self._safe_float(row.get('escore_d') or row.get('EscoreD')),
                'escore_i': self._safe_float(row.get('escore_i') or row.get('EscoreI')),
                'escore_s': self._safe_float(row.get('escore_s') or row.get('EscoreS')),
                'escore_c': self._safe_float(row.get('escore_c') or row.get('EscoreC')),
                'escore_im': self._safe_float(row.get('escore_im') or row.get('EscoreIM'))
            }

            self.supabase.table('disc_likert_map').upsert(
                data,
                on_conflict='version_ref,resposta'
            ).execute()
            count += 1

        logger.info(f"Imported {count} Likert mappings")
        return count

    def _import_rules(self, version_ref: str, rows: List[Dict]) -> int:
        """Import scoring rules."""
        count = 0

        for row in rows:
            regra_id = row.get('regra_id') or row.get('RegraID') or row.get('ID')
            if not regra_id:
                continue

            descricao_pt = row.get('descricao_pt') or row.get('Descricao') or ''
            formula_pt = row.get('formula_pt') or row.get('Formula') or ''

            data = {
                'version_ref': version_ref,
                'regra_id': str(regra_id),
                'descricao_pt': descricao_pt,
                'descricao_en': self._translate_if_missing(descricao_pt, row.get('descricao_en')),
                'formula_pt': formula_pt,
                'formula_en': self._translate_if_missing(formula_pt, row.get('formula_en')),
                'aplicacao': row.get('aplicacao') or row.get('Aplicacao')
            }

            self.supabase.table('disc_rules').upsert(
                data,
                on_conflict='version_ref,regra_id'
            ).execute()
            count += 1

        logger.info(f"Imported {count} rules")
        return count

    def _import_thresholds(self, version_ref: str, rows: List[Dict]) -> int:
        """Import interpretation thresholds (4 expected - one per factor)."""
        count = 0

        for row in rows:
            fator = row.get('fator') or row.get('Fator')
            if not fator or fator not in ['D', 'I', 'S', 'C']:
                continue

            nota_pt = row.get('nota_pt') or row.get('Nota') or ''

            data = {
                'version_ref': version_ref,
                'fator': fator,
                'baixo': self._safe_float(row.get('baixo') or row.get('Baixo')),
                'medio': self._safe_float(row.get('medio') or row.get('Medio')),
                'alto': self._safe_float(row.get('alto') or row.get('Alto')),
                'nota_pt': nota_pt,
                'nota_en': self._translate_if_missing(nota_pt, row.get('nota_en'))
            }

            self.supabase.table('disc_thresholds').upsert(
                data,
                on_conflict='version_ref,fator'
            ).execute()
            count += 1

        logger.info(f"Imported {count} thresholds")
        return count

    def _import_report_templates(self, version_ref: str, rows: List[Dict]) -> int:
        """Import report templates."""
        count = 0

        for row in rows:
            secao_key = row.get('secao_key') or row.get('SecaoKey') or row.get('Key')
            if not secao_key:
                continue

            secao_pt = row.get('secao_pt') or row.get('Secao') or ''
            texto_pt = row.get('texto_pt') or row.get('Texto') or ''
            bullets_pt = row.get('bullets_pt') or row.get('Bullets') or ''

            data = {
                'version_ref': version_ref,
                'secao_key': str(secao_key),
                'secao_pt': secao_pt,
                'secao_en': self._translate_if_missing(secao_pt, row.get('secao_en')),
                'condicao': row.get('condicao') or row.get('Condicao'),
                'texto_pt': texto_pt,
                'texto_en': self._translate_if_missing(texto_pt, row.get('texto_en')),
                'bullets_pt': bullets_pt,
                'bullets_en': self._translate_if_missing(bullets_pt, row.get('bullets_en')),
                'metricas': row.get('metricas') or row.get('Metricas'),
                'placeholders': row.get('placeholders') or row.get('Placeholders')
            }

            self.supabase.table('disc_report_templates').upsert(
                data,
                on_conflict='version_ref,secao_key'
            ).execute()
            count += 1

        logger.info(f"Imported {count} report templates")
        return count

    def _import_words(self, version_ref: str, rows: List[Dict]) -> int:
        """Import descriptive words (12 expected - 4 factors x 3 intensities)."""
        count = 0

        for row in rows:
            fator = row.get('fator') or row.get('Fator')
            intensidade = row.get('intensidade') or row.get('Intensidade')

            if not fator or not intensidade:
                continue

            lista_pt = row.get('lista_pt') or row.get('Lista') or ''

            data = {
                'version_ref': version_ref,
                'fator': fator,
                'intensidade': intensidade,
                'lista_pt': lista_pt,
                'lista_en': self._translate_if_missing(lista_pt, row.get('lista_en'))
            }

            self.supabase.table('disc_words').upsert(
                data,
                on_conflict='version_ref,fator,intensidade'
            ).execute()
            count += 1

        logger.info(f"Imported {count} word lists")
        return count

    def _import_interview(self, version_ref: str, rows: List[Dict]) -> int:
        """Import interview guide questions."""
        count = 0

        for row in rows:
            eixo = row.get('eixo') or row.get('Eixo')
            pergunta_pt = row.get('pergunta_pt') or row.get('Pergunta') or ''

            if not pergunta_pt:
                continue

            observar_pt = row.get('observar_pt') or row.get('Observar') or ''
            follow_pt = row.get('follow_pt') or row.get('Follow') or ''

            data = {
                'version_ref': version_ref,
                'eixo': eixo,
                'pergunta_pt': pergunta_pt,
                'pergunta_en': self._translate_if_missing(pergunta_pt, row.get('pergunta_en')),
                'observar_pt': observar_pt,
                'observar_en': self._translate_if_missing(observar_pt, row.get('observar_en')),
                'follow_pt': follow_pt,
                'follow_en': self._translate_if_missing(follow_pt, row.get('follow_en'))
            }

            self.supabase.table('disc_interview').insert(data).execute()
            count += 1

        logger.info(f"Imported {count} interview questions")
        return count

    def _import_quality_checks(self, version_ref: str, rows: List[Dict]) -> int:
        """Import quality validation checks."""
        count = 0

        for row in rows:
            checagem_pt = row.get('checagem_pt') or row.get('Checagem') or ''
            if not checagem_pt:
                continue

            criterio_pt = row.get('criterio_pt') or row.get('Criterio') or ''
            acao_pt = row.get('acao_pt') or row.get('Acao') or ''

            data = {
                'version_ref': version_ref,
                'checagem_pt': checagem_pt,
                'checagem_en': self._translate_if_missing(checagem_pt, row.get('checagem_en')),
                'criterio_pt': criterio_pt,
                'criterio_en': self._translate_if_missing(criterio_pt, row.get('criterio_en')),
                'acao_pt': acao_pt,
                'acao_en': self._translate_if_missing(acao_pt, row.get('acao_en'))
            }

            self.supabase.table('disc_quality_checks').insert(data).execute()
            count += 1

        logger.info(f"Imported {count} quality checks")
        return count

    def _import_ethics(self, version_ref: str, rows: List[Dict]) -> int:
        """Import ethical principles."""
        count = 0

        for row in rows:
            principio_pt = row.get('principio_pt') or row.get('Principio') or ''
            if not principio_pt:
                continue

            como_aplicamos_pt = row.get('como_aplicamos_pt') or row.get('ComoAplicamos') or ''

            data = {
                'version_ref': version_ref,
                'principio_pt': principio_pt,
                'principio_en': self._translate_if_missing(principio_pt, row.get('principio_en')),
                'como_aplicamos_pt': como_aplicamos_pt,
                'como_aplicamos_en': self._translate_if_missing(
                    como_aplicamos_pt,
                    row.get('como_aplicamos_en')
                )
            }

            self.supabase.table('disc_ethics').insert(data).execute()
            count += 1

        logger.info(f"Imported {count} ethical principles")
        return count
